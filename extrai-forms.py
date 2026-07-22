# -*- coding: utf-8 -*-
"""
Importa as respostas do Formulário de Inspeção de Segurança (Microsoft Forms /
Google Forms) para o controle de NCs. Lê os .xlsx da pasta forms/ e gera um
.txt no formato do funil (exports/) — o nc-mina.js processa sem duplicar.

Fluxo: Forms > Respostas > Abrir no Excel > salvar em forms/ >
       python extrai-forms.py > rodar.bat

Precisa de: pip install openpyxl
"""
import os, re, sys, unicodedata, datetime
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

DIR = os.path.dirname(os.path.abspath(__file__))
DIR_FORMS = os.path.join(DIR, "forms")
DIR_EXPORTS = os.path.join(DIR, "exports")

# O dashboard usa os MESMOS nomes oficiais do formulário — aqui só
# garantimos a grafia canônica (acentos/maiúsculas). Um local novo no
# formulário entra com o nome original.
AREAS_OFICIAIS = [
    "Britagem 1", "Britagem 1A", "Concentração (Planta 3 + GX600)", "Filtragem",
    "PEC", "Planta 4", "Área de Lavra",
    "Banheiro/Vestiário Cedro ADM", "Banheiro/Vestiário Cedro Operacional",
    "Banheiro/Vestiário Cedro Contratadas",
    "Pátio Balança", "Balança", "Oficina AJPM", "Oficina MPC",
    "Restaurante", "Centro Administrativo", "Portaria",
]
# (DE_PARA_AREAS é montado logo abaixo, depois da função norm)

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).lower().strip()

DE_PARA_AREAS = {norm(a): a for a in AREAS_OFICIAIS}

def fdata(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    s = str(v or "").strip()
    m = re.search(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})", s)
    if m:
        d, mo, a = m.groups()
        a = a if len(a) == 4 else "20" + a
        return f"{int(d):02d}/{int(mo):02d}/{a}"
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)  # formato ISO
    if m:
        a, mo, d = m.groups()
        return f"{int(d):02d}/{int(mo):02d}/{a}"
    return ""

def fhora(v):
    m = re.search(r"(\d{1,2})[:h](\d{2})", str(v or ""))
    return f"{int(m.group(1)):02d}:{m.group(2)}" if m else "08:00"

# reconhece as colunas pelo título da pergunta (tolerante a variações)
COLUNAS = [
    ("data",        ["data da inspecao", "data da inspeção"]),
    ("hora",        ["hora aproximada", "hora da inspecao"]),
    ("responsavel", ["responsavel(is) pela inspecao", "responsaveis pela inspecao", "responsavel pela inspecao"]),
    ("local",       ["local / ponto inspecionado", "local/ponto", "local inspecionado"]),
    ("outros",      ["se marcou \"outros\"", "especifique o local"]),
    ("tipoInsp",    ["tipo de inspecao"]),
    ("houveNC",     ["houve nao conformidade"]),
    ("tipoOcorr",   ["tipo de ocorrencia"]),
    ("risco",       ["classificacao de risco"]),
    ("gravidade",   ["potencial de gravidade"]),
    ("descricao",   ["descricao da nao conformidade"]),
    ("acaoImediata",["acao imediata"]),
    ("acaoCorretiva",["acao corretiva", "plano de acao"]),
    ("respTrat",    ["responsavel pela tratativa"]),
    ("empresa",     ["empresa responsavel", "empresa"]),
    ("prazo",       ["prazo para adequacao", "prazo"]),
    ("status",      ["status da tratativa"]),
    ("dataConcl",   ["data de conclusao"]),
    ("validacao",   ["responsavel pela verificacao", "validacao"]),
    ("empresaOutro",["qual empresa"]),
    ("foto",        ["registro fotografico", "foto", "anexo", "imagem"]),
]

def mapear_colunas(header):
    idx = {}
    hn = [norm(h) for h in header]
    for campo, chaves in COLUNAS:
        for j, h in enumerate(hn):
            if h and any(c in h for c in chaves):
                idx[campo] = j
                break
    return idx

# blocos de NC extra na mesma inspeção (perguntas "NC 2 — ..." / "NC 3 — ...")
def mapear_blocos_extra(header):
    hn = [norm(h) for h in header]
    blocos = []
    for prefixo in ("nc 2", "nc 3"):
        b = {}
        for j, h in enumerate(hn):
            if not h.startswith(prefixo):
                continue
            if "descricao" in h: b["descricao"] = j
            elif "risco" in h: b["risco"] = j
            elif "tratativa" in h and "status" not in h: b["respTrat"] = j
            elif "prazo" in h: b["prazo"] = j
            elif "status" in h: b["status"] = j
        if "descricao" in b:
            blocos.append(b)
    return blocos

def processa_tratativa(caminho, linhas_saida):
    """Planilha do Formulário 2 (Tratativa de NC): muda status pelo número."""
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    header, idx = None, {}
    n = 0
    for row in ws.iter_rows(values_only=True):
        if header is None:
            if row and any(v for v in row):
                header = [norm(h) for h in row]
                for j, h in enumerate(header):
                    if "numero da nc" in h: idx["numero"] = j
                    elif "novo status" in h or h == "status": idx["status"] = j
                    elif "observacao" in h or "evidencia" in h: idx["obs"] = j
                    elif "responsavel pela atualizacao" in h or "responsavel" in h: idx.setdefault("resp", j)
                    elif "data da atualizacao" in h or h.startswith("data"): idx.setdefault("data", j)
            continue
        get = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
        numero = str(get("numero") or "").strip().upper()
        m = re.search(r"(\d+)", numero)
        if not m:
            continue
        numero = "NC-" + m.group(1).zfill(4)
        status = str(get("status") or "").strip() or "Em andamento"
        obs = str(get("obs") or "").strip()
        resp = str(get("resp") or "").strip() or "Tratativa Forms"
        data = fdata(get("data")) or datetime.date.today().strftime("%d/%m/%Y")
        linhas_saida.append(
            f"{data} 12:00 - {resp}: #ACOMPANHAMENTO\nNC: {numero}\nStatus: {status}"
            + (f"\nObservação: {obs} (tratativa via Forms)" if obs else "\nObservação: Tratativa via Forms"))
        n += 1
    wb.close()
    return n

def eh_planilha_tratativa(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    for row in ws.iter_rows(max_row=3, values_only=True):
        if row and any("numero da nc" in norm(v) for v in row if v):
            wb.close()
            return True
    wb.close()
    return False

def processa_xlsx(caminho, linhas_saida):
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    linhas = ws.iter_rows(values_only=True)
    header = None
    idx = {}
    blocos_extra = []
    n_nc = n_visita = 0
    for row in linhas:
        if header is None:
            if row and any(v for v in row):
                header = list(row)
                idx = mapear_colunas(header)
                blocos_extra = mapear_blocos_extra(header)
                if "local" not in idx or "data" not in idx:
                    print(f"  aviso: {os.path.basename(caminho)} não parece ser a planilha do Forms (faltam colunas Data/Local)")
                    wb.close()
                    return 0, 0
            continue
        get = lambda c: (row[idx[c]] if c in idx and idx[c] < len(row) else None)
        data = fdata(get("data"))
        if not data:
            continue
        hora = fhora(get("hora"))
        # vários responsáveis podem vir juntos (Forms separa com ";")
        resp_raw = str(get("responsavel") or "").strip()
        tecnicos = [t.strip() for t in re.split(r"[;,]", resp_raw) if t.strip()] or ["(sem responsável)"]
        tecnico = tecnicos[0]

        local_raw = str(get("local") or "").strip()
        if norm(local_raw).startswith("outros"):
            local_raw = str(get("outros") or "").strip() or "Outros"
        area = DE_PARA_AREAS.get(norm(local_raw), local_raw)

        descricao = str(get("descricao") or "").strip()
        houve = norm(get("houveNC"))
        eh_nc = descricao and not houve.startswith("nao")

        if not eh_nc:
            # inspeção sem desvio: registra a visita (controle de frequência)
            linhas_saida.append(f"{data} {hora} - {tecnico}: #INSPECAO {area} - tudo ok (via Forms{', com ' + ', '.join(tecnicos[1:]) if len(tecnicos) > 1 else ''})")
            n_visita += 1
            continue

        obs_partes = []
        if get("tipoOcorr"): obs_partes.append(f"Tipo: {str(get('tipoOcorr')).strip()}")
        if get("gravidade"): obs_partes.append(f"Gravidade potencial: {str(get('gravidade')).strip()}")
        if get("acaoImediata") and norm(get("acaoImediata")) not in ("", "nao aplicavel", "n/a", "na"):
            obs_partes.append(f"Ação imediata: {str(get('acaoImediata')).strip()}")
        if get("acaoCorretiva"): obs_partes.append(f"Plano de ação: {str(get('acaoCorretiva')).strip()}")
        empresa = str(get("empresa") or "").strip()
        if norm(empresa) == "outro" and get("empresaOutro"):
            empresa = str(get("empresaOutro")).strip()  # nome escrito quando marca "Outro"
        if empresa: obs_partes.append(f"Empresa: {empresa}")
        if len(tecnicos) > 1: obs_partes.append(f"Equipe: {', '.join(tecnicos)}")

        corpo = [f"{data} {hora} - {tecnico}: #NC",
                 f"Área: {area}",
                 f"Não Conformidade: {descricao} (via Forms)"]
        if get("risco"): corpo.append(f"Classificação: {str(get('risco')).strip()}")
        if get("respTrat"): corpo.append(f"Responsável: {str(get('respTrat')).strip()}")
        prazo = fdata(get("prazo"))
        if prazo: corpo.append(f"Prazo: {prazo}")
        if obs_partes: corpo.append("Observação: " + " | ".join(obs_partes))
        foto = str(get("foto") or "").strip()
        if foto.startswith("http"):
            corpo.append(f"Foto: {foto}")
        linhas_saida.append("\n".join(corpo))
        n_nc += 1

        # status já resolvido/em andamento no próprio formulário
        status = norm(get("status"))
        if status.startswith("conclu"):
            dc = fdata(get("dataConcl")) or data
            valid = str(get("validacao") or "").strip()
            linhas_saida.append(
                f"{dc} 17:00 - {tecnico}: #ACOMPANHAMENTO\nÁrea: {area}\nNC: {descricao[:90]}\n"
                f"Status: Concluído\nObservação: Concluída via Forms{(' · verificado por ' + valid) if valid else ''}")
        elif "andamento" in status:
            linhas_saida.append(
                f"{data} {hora} - {tecnico}: #ACOMPANHAMENTO\nÁrea: {area}\nNC: {descricao[:90]}\n"
                f"Status: Em andamento\nObservação: Status informado no Forms")

        # NCs extras da MESMA inspeção (blocos "NC 2 — ..." / "NC 3 — ...")
        for b in blocos_extra:
            d_extra = str(row[b["descricao"]] or "").strip() if b["descricao"] < len(row) else ""
            if not d_extra:
                continue
            gb = lambda c: (row[b[c]] if c in b and b[c] < len(row) else None)
            corpo2 = [f"{data} {hora} - {tecnico}: #NC",
                      f"Área: {area}",
                      f"Não Conformidade: {d_extra} (via Forms)"]
            if gb("risco"): corpo2.append(f"Classificação: {str(gb('risco')).strip()}")
            if gb("respTrat"): corpo2.append(f"Responsável: {str(gb('respTrat')).strip()}")
            p2 = fdata(gb("prazo"))
            if p2: corpo2.append(f"Prazo: {p2}")
            linhas_saida.append("\n".join(corpo2))
            n_nc += 1
            st2 = norm(gb("status"))
            if st2.startswith("conclu"):
                linhas_saida.append(
                    f"{data} 17:00 - {tecnico}: #ACOMPANHAMENTO\nÁrea: {area}\nNC: {d_extra[:90]}\n"
                    f"Status: Concluído\nObservação: Concluída via Forms")
            elif "andamento" in st2:
                linhas_saida.append(
                    f"{data} {hora} - {tecnico}: #ACOMPANHAMENTO\nÁrea: {area}\nNC: {d_extra[:90]}\n"
                    f"Status: Em andamento\nObservação: Status informado no Forms")
    wb.close()
    return n_nc, n_visita

os.makedirs(DIR_FORMS, exist_ok=True)
os.makedirs(DIR_EXPORTS, exist_ok=True)
arquivos = [f for f in os.listdir(DIR_FORMS) if f.lower().endswith(".xlsx") and not f.startswith("~$")]
if not arquivos:
    print(f"Nenhum .xlsx em {DIR_FORMS} — baixe as respostas do Forms (Abrir no Excel) e salve lá.")
tot_nc = tot_v = 0
saida = []
for f in arquivos:
    print(f"Lendo: {f}")
    caminho = os.path.join(DIR_FORMS, f)
    if eh_planilha_tratativa(caminho):
        n = processa_tratativa(caminho, saida)
        print(f"  planilha de TRATATIVA: {n} atualização(ões) de status")
        continue
    n, v = processa_xlsx(caminho, saida)
    print(f"  {n} NC(s), {v} visita(s) sem desvio")
    tot_nc += n; tot_v += v
if saida:
    arq = os.path.join(DIR_EXPORTS, "forms_respostas.txt")
    with open(arq, "w", encoding="utf-8") as fp:
        fp.write("\n".join(saida) + "\n")
    print(f"\nGerado: {arq}")
print(f"Total: {tot_nc} NC(s) e {tot_v} visita(s). Agora rode o rodar.bat.")
