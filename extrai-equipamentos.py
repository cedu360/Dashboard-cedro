# -*- coding: utf-8 -*-
"""
Extrai o cadastro de equipamentos da base MANUTENÇÃO GERENCIAL e gera
equipamentos.json na pasta do projeto. Rode de novo sempre que a base mudar:
    python extrai-equipamentos.py
Precisa de: pip install openpyxl
"""
import os, re, json, sys, unicodedata
import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:\Users\pc\OneDrive\Documentos\MANUTENÇÃO GERENCIAL"
SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "equipamentos.json")

# pastas que contêm planilhas de equipamento (1 arquivo = 1 equipamento, ou abas = equipamentos)
PASTAS_EQUIP = ["BRITAGEM", "CONCENTRAÇÃO", "FILTRAGEM", "EQUIPAMENTOS"]
# subpastas/arquivos que NÃO são equipamentos (estoque, containers, ferramentas, materiais)
IGNORAR = ["CONTAINER", "FERRAMENTARIA", "ARMAZENAGEM", "MISCELANEA", "COMPONENTES", "CONTROLE DE BORRACHAS"]

LOCAL_PARA_AREA = {
    "BRITAGEM": "Britagem 1",
    "CONCENTRACAO": "Concentração (Planta 3 + GX600)",
    "FILTRAGEM": "Filtragem",
}

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).upper().strip()

def area_do_local(local, pasta_raiz):
    n = norm(local)
    for chave, area in LOCAL_PARA_AREA.items():
        if chave in n:
            return area
    return LOCAL_PARA_AREA.get(norm(pasta_raiz), None)

CONDICOES_ATENCAO = ("MEIA VIDA", "DESGAST", "RUIM", "TROCAR", "DANIFICAD", "INOPERANTE", "PARAD")

def canon_tag(tag):
    # normaliza grafias: "BP - 01 /BP- 01R" e "BP-01 & BP-01R" => mesma chave;
    # remove nomes de fabricante entre parênteses: "PN-01 (AÇOITA)" => PN01
    s = re.sub(r"\([^)]*\)", "", norm(tag))
    return re.sub(r"[^A-Z0-9]", "", s)

equipamentos = {}   # canon => registro final (o arquivo com MAIS componentes vence; sem soma dupla)
locais_votos = {}   # canon => {local: contagem} para decidir a área por maioria

def processa_aba(ws, nome_arquivo, pasta_raiz, acumulador):
    # acha a linha de cabeçalho (a que contém "TAG")
    header, header_row = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True), 1):
        vals = [norm(v) for v in row]
        if "TAG" in vals:
            header, header_row = vals, i
            break
    if not header:
        return
    def col(*nomes):
        for n in nomes:
            for j, h in enumerate(header):
                if h and n in h:
                    return j
        return None
    c_tag, c_local, c_modelo = col("TAG"), col("LOCAL"), col("MODELO")
    c_peca, c_cond = col("PECAS", "PEÇAS", "APLICACAO", "DESCRICAO"), col("CONDICAO")
    if c_tag is None:
        return

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        tag = row[c_tag] if c_tag < len(row) else None
        if not tag or norm(tag) in ("TAG", ""):
            continue
        tag = str(tag).strip()
        chave = canon_tag(tag)
        if not chave:
            continue
        local = row[c_local] if c_local is not None and c_local < len(row) else ""
        a = area_do_local(local, pasta_raiz)  # LOCAL da planilha; se vazio, cai na pasta
        if a:
            votos = locais_votos.setdefault(chave, {})
            # voto da coluna LOCAL vale mais que o da pasta
            votos[a] = votos.get(a, 0) + (2 if local else 1)
        e = acumulador.get(chave)
        if not e:
            e = acumulador[chave] = {
                "tag": tag,
                "nome": os.path.splitext(nome_arquivo)[0].strip(),
                "modelo": "",
                "componentes": 0,
                "atencao": 0,          # componentes com condição ruim/meia vida
                "condicoes": {},
            }
        modelo = row[c_modelo] if c_modelo is not None and c_modelo < len(row) else ""
        if modelo and not e["modelo"]:
            e["modelo"] = str(modelo).strip()
        e["componentes"] += 1
        cond = norm(row[c_cond]) if c_cond is not None and c_cond < len(row) and row[c_cond] else ""
        if cond:
            e["condicoes"][cond] = e["condicoes"].get(cond, 0) + 1
            if any(p in cond for p in CONDICOES_ATENCAO):
                e["atencao"] += 1

total_arq = 0
for pasta in PASTAS_EQUIP:
    raiz = os.path.join(BASE, pasta)
    if not os.path.isdir(raiz):
        continue
    for dirpath, dirnames, filenames in os.walk(raiz):
        if any(ig in norm(dirpath) for ig in IGNORAR):
            continue
        for f in filenames:
            if not f.lower().endswith(".xlsx") or f.startswith("~$"):
                continue
            if any(ig in norm(f) for ig in IGNORAR):
                continue
            caminho = os.path.join(dirpath, f)
            try:
                wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
                por_arquivo = {}
                for ws in wb.worksheets:
                    processa_aba(ws, f, pasta, por_arquivo)
                wb.close()
                total_arq += 1
                # mesmo equipamento em vários arquivos (ex.: BP-01.xlsx e BOMBAS COMPLETO.xlsx):
                # fica a versão com MAIS componentes — nunca soma duplicado
                for chave, e in por_arquivo.items():
                    atual = equipamentos.get(chave)
                    if not atual or e["componentes"] > atual["componentes"]:
                        equipamentos[chave] = e
            except Exception as ex:
                print(f"  aviso: não li {f}: {ex}")

# define a área pela maioria dos registros de LOCAL
for chave, e in equipamentos.items():
    votos = locais_votos.get(chave, {})
    e["area"] = max(votos, key=votos.get) if votos else None

lista = sorted(equipamentos.values(), key=lambda e: (e["area"] or "z", e["tag"]))
with open(SAIDA, "w", encoding="utf-8") as fp:
    json.dump(lista, fp, ensure_ascii=False, indent=1)

print(f"Arquivos lidos: {total_arq}")
print(f"Equipamentos: {len(lista)}")
por_area = {}
for e in lista:
    por_area[e["area"] or "(sem área)"] = por_area.get(e["area"] or "(sem área)", 0) + 1
for a, n in sorted(por_area.items()):
    print(f"  {a}: {n}")
print(f"Gerado: {SAIDA}")
