# -*- coding: utf-8 -*-
"""
Extrai o fluxo de solicitações de manutenção/compras da base MANUTENÇÃO GERENCIAL
(planilhas ACOMPANHAMENTO *.xlsm) e gera solicitacoes.json na pasta do projeto.
Rode de novo sempre que as planilhas mudarem:
    python extrai-solicitacoes.py
Precisa de: pip install openpyxl
"""
import os, json, sys, unicodedata, datetime, warnings
import openpyxl

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8")

BASE = r"C:\Users\pc\OneDrive\Documentos\MANUTENÇÃO GERENCIAL\ACOMPANHAMENTO DAS SOLICITAÇÕES"
SAIDA = os.path.join(os.path.dirname(os.path.abspath(__file__)), "solicitacoes.json")

ARQUIVOS = {
    "ACOMPANHAMENTO BRITAGEM.xlsm": "Britagem 1",
    "ACOMPANHAMERNTO CONCENTRAÇÃO.xlsm": "Concentração (Planta 3 + GX600)",
    "ACOMPANHAMERNTO ELETRICA.xlsm": "Elétrica",
}
ABAS = {"OM em Tratamento": "Em aberto", "Processos Concluídos": "Concluída"}

def norm(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if not unicodedata.combining(c)).upper().strip()

def fdata(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip() if v else ""

def le_aba(ws, setor, status, saida):
    # acha a linha de cabeçalho (contém "SOLICITANTE")
    header, header_row = None, None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        vals = [norm(v) for v in row]
        if any("SOLICITANTE" in v for v in vals if v):
            header, header_row = vals, i
            break
    if not header:
        return 0
    def col(*nomes):
        for n in nomes:
            for j, h in enumerate(header):
                if h and n in h:
                    return j
        return None
    c_sol = col("SOLICITANTE")
    c_dt = col("DT DE SOLICITACAO", "DATA DE SOLICITACAO")
    c_txt = col("TXT", "BREVE", "TAG DO EQUIP")
    c_tag = col("TAG")
    if c_tag == c_txt:
        c_tag = None
    c_lib = col("LIBERACAO")
    c_lead = col("LEAD TIME")
    c_ana = col("ANALISTA")
    c_ord = col("ORDEM")
    c_oc = col("OC")
    n = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        get = lambda c: (row[c] if c is not None and c < len(row) else None)
        solicitante = str(get(c_sol) or "").strip()
        texto = str(get(c_txt) or "").strip()
        tag = str(get(c_tag) or "").strip()
        if not solicitante and not texto:
            continue
        if norm(solicitante) in ("SOLICITANTE",) or norm(texto).startswith("TXT"):
            continue
        lead = get(c_lead)
        try:
            lead = int(float(lead)) if lead not in (None, "") else None
        except (TypeError, ValueError):
            lead = None
        if lead is not None and (lead < 0 or lead > 3650):
            lead = None  # célula com data serial do Excel ou lixo, não é lead time
        saida.append({
            "setor": setor,
            "status": status,
            "solicitante": solicitante,
            "data": fdata(get(c_dt)),
            "texto": texto,
            "tag": tag,
            "dataLiberacao": fdata(get(c_lib)),
            "leadTime": lead,
            "analista": str(get(c_ana) or "").strip().replace("#REF!", ""),
            "ordem": str(get(c_ord) or "").strip(),
            "oc": str(get(c_oc) or "").strip(),
        })
        n += 1
    return n

todas = []
for arq, setor in ARQUIVOS.items():
    caminho = os.path.join(BASE, arq)
    if not os.path.isfile(caminho):
        print(f"  aviso: não achei {arq}")
        continue
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    for nome_aba, status in ABAS.items():
        if nome_aba in wb.sheetnames:
            n = le_aba(wb[nome_aba], setor, status, todas)
            print(f"  {setor} / {nome_aba}: {n}")
    wb.close()

with open(SAIDA, "w", encoding="utf-8") as fp:
    json.dump(todas, fp, ensure_ascii=False, indent=0)

abertas = [s for s in todas if s["status"] == "Em aberto"]
print(f"\nTotal: {len(todas)} solicitações ({len(abertas)} em aberto)")
print(f"Gerado: {SAIDA}")
